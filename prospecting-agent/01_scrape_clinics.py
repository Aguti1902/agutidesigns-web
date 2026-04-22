"""
AGENTE DE SCRAPING — Clínicas Dentales
=======================================
Busca clínicas dentales en Google Maps (Barcelona, Madrid, Sevilla, Valencia),
entra en la web de cada una para extraer el email de contacto,
y guarda todo en un Excel listo para usar.

Uso:
  python 01_scrape_clinics.py

Resultado:
  clinicas_dentales.xlsx
"""

import os
import re
import time
import logging
import googlemaps
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from tqdm import tqdm

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# ── Configuración ──────────────────────────────────────────────────────────────

API_KEY   = os.getenv("GOOGLE_PLACES_API_KEY", "")
CITIES    = ["Barcelona", "Madrid", "Sevilla", "Valencia"]
QUERY     = "clínica dental"
OUTPUT    = "clinicas_dentales.xlsx"

# Palabras que suelen indicar una página de contacto
CONTACT_HINTS = ["contacto", "contact", "contactenos", "contacta", "about", "sobre"]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "es-ES,es;q=0.9",
}

EMAIL_RE = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", re.IGNORECASE
)
# Dominios que son redes sociales o placeholders — ignorar
SKIP_EMAIL_DOMAINS = {
    "ejemplo.com", "example.com", "sentry.io", "wixpress.com",
    "squarespace.com", "wordpress.com", "google.com",
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def clean_email(email: str) -> str | None:
    """Limpia y valida un email encontrado."""
    email = email.strip().lower()
    domain = email.split("@")[-1]
    if domain in SKIP_EMAIL_DOMAINS:
        return None
    if len(email) > 80 or "." not in domain:
        return None
    return email


def extract_emails_from_html(html: str) -> list[str]:
    """Extrae todos los emails únicos de un bloque HTML."""
    found = EMAIL_RE.findall(html)
    clean = []
    for e in found:
        c = clean_email(e)
        if c and c not in clean:
            clean.append(c)
    return clean


def find_contact_url(soup: BeautifulSoup, base_url: str) -> str | None:
    """Busca un enlace de 'Contacto' en la página."""
    for a in soup.find_all("a", href=True):
        href = a["href"].lower()
        text = a.get_text(strip=True).lower()
        for hint in CONTACT_HINTS:
            if hint in href or hint in text:
                full = href if href.startswith("http") else base_url.rstrip("/") + "/" + href.lstrip("/")
                return full
    return None


def scrape_email(website: str, timeout: int = 8) -> str | None:
    """
    Intenta extraer un email de la web de la clínica:
    1. Página principal
    2. Página de contacto (si la encuentra)
    """
    if not website:
        return None

    base = website.rstrip("/")
    emails: list[str] = []

    try:
        resp = requests.get(base, headers=HEADERS, timeout=timeout, allow_redirects=True)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")
        emails = extract_emails_from_html(resp.text)

        if not emails:
            # Busca página de contacto
            contact_url = find_contact_url(soup, base)
            if contact_url and contact_url != base:
                time.sleep(0.5)
                cr = requests.get(contact_url, headers=HEADERS, timeout=timeout, allow_redirects=True)
                emails = extract_emails_from_html(cr.text)

    except Exception as exc:
        log.debug("Error scraping %s: %s", website, exc)

    return emails[0] if emails else None


# ── Google Places API ──────────────────────────────────────────────────────────

def search_clinics_in_city(gmaps: googlemaps.Client, city: str) -> list[dict]:
    """
    Busca clínicas dentales en una ciudad y devuelve todos los resultados
    paginando hasta 3 páginas (máx. ~60 resultados por ciudad).
    """
    query = f"{QUERY} en {city}"
    log.info("Buscando: %s", query)
    results = []

    try:
        response = gmaps.places(query=query, language="es")
        results.extend(response.get("results", []))

        for _ in range(2):  # Hasta 2 páginas extra
            token = response.get("next_page_token")
            if not token:
                break
            time.sleep(2.5)  # Google requiere esperar antes de usar el token
            response = gmaps.places(query=query, language="es", page_token=token)
            results.extend(response.get("results", []))

    except Exception as exc:
        log.error("Error buscando en %s: %s", city, exc)

    log.info("  → %d clínicas encontradas en %s", len(results), city)
    return results


def get_place_details(gmaps: googlemaps.Client, place_id: str) -> dict:
    """
    Obtiene los detalles completos de un lugar:
    teléfono, web, dirección formateada, horarios, etc.
    """
    try:
        fields = [
            "name", "formatted_address", "formatted_phone_number",
            "website", "rating", "user_ratings_total", "url",
        ]
        detail = gmaps.place(place_id=place_id, fields=fields, language="es")
        return detail.get("result", {})
    except Exception as exc:
        log.debug("Error obteniendo detalles de %s: %s", place_id, exc)
        return {}


# ── Excel ──────────────────────────────────────────────────────────────────────

BLUE_FILL  = PatternFill("solid", fgColor="0047FF")
GRAY_FILL  = PatternFill("solid", fgColor="F5F5F5")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin", color="E0E0E0"),
    top=Side(style="thin", color="E0E0E0"),
    bottom=Side(style="thin", color="E0E0E0"),
)

COLUMNS = [
    ("Ciudad",          20),
    ("Nombre",          35),
    ("Dirección",       45),
    ("Teléfono",        18),
    ("Email",           32),
    ("Web",             40),
    ("Valoración",      12),
    ("Nº Reseñas",      12),
    ("Email enviado",   14),
    ("Notas",           30),
]


def build_excel(rows: list[dict], output: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Clínicas Dentales"

    # Cabecera
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # Datos
    for row_idx, row in enumerate(rows, start=2):
        fill = GRAY_FILL if row_idx % 2 == 0 else WHITE_FILL
        values = [
            row.get("ciudad", ""),
            row.get("nombre", ""),
            row.get("direccion", ""),
            row.get("telefono", ""),
            row.get("email", ""),
            row.get("web", ""),
            row.get("rating", ""),
            row.get("reviews", ""),
            "No",   # Email enviado
            "",     # Notas
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if col_idx == 5 and value and "@" in str(value):  # Email en azul
                cell.font = Font(color="0047FF")

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    wb.save(output)
    log.info("Excel guardado en: %s (%d filas)", output, len(rows))


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    if not API_KEY:
        log.error("Falta GOOGLE_PLACES_API_KEY en el archivo .env")
        return

    gmaps = googlemaps.Client(key=API_KEY)
    all_rows: list[dict] = []
    seen_place_ids: set[str] = set()

    for city in CITIES:
        places = search_clinics_in_city(gmaps, city)

        for place in tqdm(places, desc=f"Procesando {city}", unit="clínica"):
            place_id = place.get("place_id", "")
            if place_id in seen_place_ids:
                continue
            seen_place_ids.add(place_id)

            # Detalles completos
            time.sleep(0.15)
            details = get_place_details(gmaps, place_id)

            name    = details.get("name") or place.get("name", "")
            address = details.get("formatted_address") or place.get("formatted_address", "")
            phone   = details.get("formatted_phone_number", "")
            web     = details.get("website", "")
            rating  = details.get("rating", "")
            reviews = details.get("user_ratings_total", "")

            # Scraping de email desde la web de la clínica
            email = ""
            if web:
                time.sleep(0.3)
                email = scrape_email(web) or ""

            all_rows.append({
                "ciudad":    city,
                "nombre":    name,
                "direccion": address,
                "telefono":  phone,
                "email":     email,
                "web":       web,
                "rating":    rating,
                "reviews":   reviews,
            })

    build_excel(all_rows, OUTPUT)

    total   = len(all_rows)
    with_email = sum(1 for r in all_rows if r["email"])
    log.info("=" * 50)
    log.info("RESUMEN FINAL")
    log.info("  Total clínicas:        %d", total)
    log.info("  Con email encontrado:  %d (%.0f%%)", with_email, with_email / total * 100 if total else 0)
    log.info("  Sin email:             %d", total - with_email)
    log.info("  Archivo: %s", OUTPUT)


if __name__ == "__main__":
    main()
