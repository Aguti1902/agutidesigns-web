"""
AGENTE DE EMAIL — Envío personalizado con IA
=============================================
Lee el Excel generado por 01_scrape_clinics.py,
usa GPT-4o para escribir un email 100% personalizado por clínica
y lo envía por SMTP.

Solo envía a clínicas con email y que NO tengan "Email enviado = Sí".
Marca cada fila como enviada en el Excel para no duplicar envíos.

Uso:
  python 02_send_emails.py                    # Envío real
  python 02_send_emails.py --dry-run          # Solo muestra los emails, no los envía
  python 02_send_emails.py --limit 10         # Envía solo los primeros N
"""

import os
import sys
import time
import smtplib
import logging
import argparse
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openai
from openpyxl import load_workbook
from dotenv import load_dotenv
from tqdm import tqdm

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# ── Config ─────────────────────────────────────────────────────────────────────

EXCEL_FILE  = "clinicas_dentales.xlsx"
SHEET_NAME  = "Clínicas Dentales"

OPENAI_KEY  = os.getenv("OPENAI_API_KEY", "")
SMTP_HOST   = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT   = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER   = os.getenv("SMTP_USER", "")
SMTP_PASS   = os.getenv("SMTP_PASS", "")

SENDER_NAME     = os.getenv("SENDER_NAME", "Alejandro")
SENDER_COMPANY  = os.getenv("SENDER_COMPANY", "Agutidesigns")
SENDER_WEBSITE  = os.getenv("SENDER_WEBSITE", "https://agutidesigns.io")
SENDER_WHATSAPP = os.getenv("SENDER_WHATSAPP", "+34625204337")

# Columnas del Excel (índice 1-based)
COL = {
    "ciudad":         1,
    "nombre":         2,
    "direccion":      3,
    "telefono":       4,
    "email":          5,
    "web":            6,
    "rating":         7,
    "reviews":        8,
    "email_enviado":  9,
    "notas":          10,
}

# Pausa entre envíos para no saturar el SMTP
SEND_DELAY_SECONDS = 3


# ── Generador de email con IA ──────────────────────────────────────────────────

SYSTEM_PROMPT = f"""Eres {SENDER_NAME}, fundador de {SENDER_COMPANY} — estudio de diseño web especializado en clínicas dentales.

Tu trabajo es escribir emails de prospección B2B personalizados, breves y directos para clínicas dentales en España.

SOBRE LO QUE OFRECEMOS:
- Webs diseñadas a medida para clínicas dentales que aparecen en el top de Google
- SEO local activo mensual para captar pacientes que buscan "dentista en [ciudad]"
- Google Ads y Meta Ads gestionados
- Asistente de IA que atiende pacientes y agenda citas automáticamente 24/7
- Panel de control con métricas de visitas, pacientes nuevos, posición SEO, conversiones
- Planes desde 149€/mes (sin permanencia)

REGLAS DE ESCRITURA:
- El email debe parecer escrito a mano, no una plantilla
- Máximo 150 palabras en el cuerpo
- Menciona el nombre de la clínica y la ciudad al menos una vez de forma natural
- Si tiene valoración alta (>4.2), menciona que tiene buenas reseñas pero que podría captar aún más pacientes con mejor visibilidad online
- Tono: cercano, profesional y directo. Nada de "Estimado/a señor/a"
- NO uses bullet points ni listas — texto fluido como un email real
- Termina SIEMPRE con: un enlace a WhatsApp y la web
- Incluye siempre al final una línea de baja: "Si no deseas recibir más emails de nuestra parte, responde con BAJA."
- Escribe en español

FORMATO DE SALIDA:
Devuelve exactamente:
ASUNTO: [asunto del email]
---
[cuerpo del email en HTML básico: usa <p> para párrafos, <a> para links, <strong> para negritas]
"""


def generate_email(clinic: dict) -> tuple[str, str]:
    """
    Genera asunto + cuerpo HTML del email personalizado para una clínica.
    Devuelve (subject, html_body).
    """
    client = openai.OpenAI(api_key=OPENAI_KEY)

    rating_text = ""
    if clinic.get("rating"):
        try:
            r = float(clinic["rating"])
            reviews = clinic.get("reviews", "")
            rating_text = f"Su valoración en Google es {r}/5 ({reviews} reseñas)."
        except ValueError:
            pass

    user_message = f"""
Escribe un email de prospección para esta clínica dental:

- Nombre: {clinic['nombre']}
- Ciudad: {clinic['ciudad']}
- Dirección: {clinic.get('direccion', 'No disponible')}
- Web actual: {clinic.get('web', 'No tiene web visible')}
- Teléfono: {clinic.get('telefono', 'No disponible')}
- {rating_text}

Mi WhatsApp: {SENDER_WHATSAPP}
Mi web: {SENDER_WEBSITE}
Mi nombre: {SENDER_NAME} ({SENDER_COMPANY})
"""

    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": user_message},
        ],
        temperature=0.85,
        max_tokens=600,
    )

    raw = response.choices[0].message.content.strip()

    # Parsear asunto y cuerpo
    if "---" in raw:
        parts = raw.split("---", 1)
        subject_line = parts[0].strip()
        body_html    = parts[1].strip()
    else:
        subject_line = f"Tu clínica dental en {clinic['ciudad']} puede captar más pacientes"
        body_html    = raw

    # Limpiar "ASUNTO: " del inicio
    subject = subject_line.replace("ASUNTO:", "").strip()

    return subject, body_html


# ── Envío de email ─────────────────────────────────────────────────────────────

EMAIL_TEMPLATE = """
<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <style>
    body {{ font-family: 'Helvetica Neue', Arial, sans-serif; background:#f9f9f9; margin:0; padding:0; }}
    .wrap {{ max-width:580px; margin:30px auto; background:#fff; border-radius:8px;
             border:1px solid #e8e8e8; overflow:hidden; }}
    .header {{ background:#0047FF; padding:24px 32px; }}
    .header h1 {{ color:#fff; font-size:18px; margin:0; letter-spacing:-0.3px; }}
    .header p  {{ color:rgba(255,255,255,0.7); font-size:12px; margin:4px 0 0; }}
    .body {{ padding:32px; color:#1B1B1B; font-size:15px; line-height:1.7; }}
    .body p {{ margin:0 0 16px; }}
    .body a {{ color:#0047FF; }}
    .cta {{ display:inline-block; margin-top:8px; background:#25D366; color:#fff !important;
            padding:12px 24px; border-radius:50px; text-decoration:none; font-weight:700;
            font-size:14px; }}
    .footer {{ background:#f5f5f5; padding:16px 32px; font-size:11px; color:#999;
               border-top:1px solid #e8e8e8; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="header">
      <h1>{sender_company}</h1>
      <p>Diseño web para clínicas dentales · {sender_website}</p>
    </div>
    <div class="body">
      {body}
      <p><a class="cta" href="https://wa.me/{wa_clean}?text=Hola+{sender_name}%2C+he+recibido+tu+email+sobre+mi+cl%C3%ADnica">
        📱 Hablar por WhatsApp
      </a></p>
    </div>
    <div class="footer">
      {sender_name} · {sender_company} · {sender_website}<br>
      Si no deseas recibir más emails de nuestra parte, responde con <strong>BAJA</strong>.
    </div>
  </div>
</body>
</html>
"""


def send_email(to_email: str, subject: str, html_body: str, dry_run: bool = False) -> bool:
    """Envía el email por SMTP. Si dry_run=True, solo lo muestra."""
    if dry_run:
        log.info("[DRY RUN] Para: %s | Asunto: %s", to_email, subject)
        print("\n" + "="*60)
        print(f"PARA:   {to_email}")
        print(f"ASUNTO: {subject}")
        print("-"*60)
        from html.parser import HTMLParser
        class Stripper(HTMLParser):
            def __init__(self):
                super().__init__()
                self.text = []
            def handle_data(self, d):
                self.text.append(d)
            def get_text(self):
                return " ".join(self.text)
        s = Stripper()
        s.feed(html_body)
        print(s.get_text()[:500])
        print("="*60 + "\n")
        return True

    wa_clean = SENDER_WHATSAPP.replace("+", "").replace(" ", "")
    full_html = EMAIL_TEMPLATE.format(
        sender_name=SENDER_NAME,
        sender_company=SENDER_COMPANY,
        sender_website=SENDER_WEBSITE,
        wa_clean=wa_clean,
        body=html_body,
    )

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = f"{SENDER_NAME} · {SENDER_COMPANY} <{SMTP_USER}>"
    msg["To"]      = to_email

    msg.attach(MIMEText(full_html, "html", "utf-8"))

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, to_email, msg.as_string())
        return True
    except Exception as exc:
        log.error("Error enviando a %s: %s", to_email, exc)
        return False


# ── Main ───────────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(description="Agente de emails para clínicas dentales")
    parser.add_argument("--dry-run", action="store_true", help="Muestra emails sin enviarlos")
    parser.add_argument("--limit",   type=int, default=0,  help="Máximo de emails a enviar (0 = todos)")
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if not OPENAI_KEY:
        log.error("Falta OPENAI_API_KEY en el archivo .env")
        sys.exit(1)

    if not args.dry_run and (not SMTP_USER or not SMTP_PASS):
        log.error("Faltan SMTP_USER o SMTP_PASS en el archivo .env")
        sys.exit(1)

    # Cargar Excel
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME]
    except FileNotFoundError:
        log.error("No se encuentra %s — ejecuta primero 01_scrape_clinics.py", EXCEL_FILE)
        sys.exit(1)

    # Recoger clínicas pendientes
    pending = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        email_cell   = row[COL["email"] - 1]
        sent_cell    = row[COL["email_enviado"] - 1]
        nombre_cell  = row[COL["nombre"] - 1]

        email_val    = str(email_cell.value or "").strip()
        sent_val     = str(sent_cell.value or "").strip().lower()

        if email_val and "@" in email_val and sent_val != "sí":
            pending.append({
                "row_idx":   email_cell.row,
                "nombre":    str(nombre_cell.value or ""),
                "ciudad":    str(row[COL["ciudad"] - 1].value or ""),
                "direccion": str(row[COL["direccion"] - 1].value or ""),
                "telefono":  str(row[COL["telefono"] - 1].value or ""),
                "email":     email_val,
                "web":       str(row[COL["web"] - 1].value or ""),
                "rating":    str(row[COL["rating"] - 1].value or ""),
                "reviews":   str(row[COL["reviews"] - 1].value or ""),
                "sent_cell": sent_cell,
                "notes_cell": row[COL["notas"] - 1],
            })

    total = len(pending)
    log.info("Clínicas pendientes de contactar: %d", total)

    if args.limit > 0:
        pending = pending[:args.limit]
        log.info("Limitando a %d emails", args.limit)

    sent_ok = 0
    sent_fail = 0

    for clinic in tqdm(pending, desc="Enviando emails", unit="email"):
        try:
            # Generar email con IA
            subject, body = generate_email(clinic)

            # Enviar
            ok = send_email(clinic["email"], subject, body, dry_run=args.dry_run)

            if ok:
                sent_ok += 1
                if not args.dry_run:
                    clinic["sent_cell"].value   = "Sí"
                    clinic["notes_cell"].value  = f"Enviado: {subject[:40]}"
            else:
                sent_fail += 1
                clinic["notes_cell"].value = "Error al enviar"

        except Exception as exc:
            log.error("Error procesando %s: %s", clinic["nombre"], exc)
            sent_fail += 1
            clinic["notes_cell"].value = f"Error IA: {str(exc)[:60]}"

        # Guardar Excel tras cada envío (por si se interrumpe)
        if not args.dry_run:
            wb.save(EXCEL_FILE)

        time.sleep(SEND_DELAY_SECONDS)

    # Resumen
    log.info("=" * 50)
    log.info("RESUMEN DE ENVÍOS")
    log.info("  Enviados correctamente: %d", sent_ok)
    log.info("  Errores:                %d", sent_fail)
    log.info("  Excel actualizado:      %s", EXCEL_FILE)


if __name__ == "__main__":
    main()
